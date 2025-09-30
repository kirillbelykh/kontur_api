import time
import json
import os
from pathlib import Path
from typing import Dict, Any, List, Optional, Tuple
import requests
import datetime
from logger import logger
import win32com.client
from win32com.client import Dispatch
import pythoncom
from dotenv import load_dotenv
import time
import socket
import base64
import os
import openpyxl  # Для создания XLS
import tempfile

load_dotenv()

# ---------------- config ----------------
BASE = os.getenv("BASE_URL")
ORGANIZATION_ID = os.getenv("ORGANIZATION_ID")
OMS_ID = os.getenv("OMS_ID")
WAREHOUSE_ID = os.getenv("WAREHOUSE_ID")
PRODUCT_GROUP = os.getenv("PRODUCT_GROUP")
RELEASE_METHOD_TYPE = os.getenv("RELEASE_METHOD_TYPE")
CIS_TYPE = os.getenv("CIS_TYPE")
FILLING_METHOD = os.getenv("FILLING_METHOD")

# debug files
LAST_SINGLE_REQ = Path("last_single_request.json")
LAST_SINGLE_RESP = Path("last_single_response.json")
LAST_MULTI_LOG = Path("last_multistep_log.json")

# CAdES / CAPICOM constants
CADES_BES = 1
CADESCOM_BASE64_TO_BINARY = 1
CAPICOM_ENCODE_BASE64 = 0
CAPICOM_AUTHENTICATED_ATTRIBUTE_SIGNING_TIME = 0
CAPICOM_CURRENT_USER_STORE = 2
CAPICOM_MY_STORE = "My"
CAPICOM_STORE_OPEN_MAXIMUM_ALLOWED = 2


# ---------- certificate utilities (pywin32 / CAdES) ----------
def find_certificate_by_thumbprint(thumbprint: Optional[str] = None):
    pythoncom.CoInitialize()
    store = win32com.client.Dispatch("CAdESCOM.Store")
    store.Open(CAPICOM_CURRENT_USER_STORE, CAPICOM_MY_STORE, CAPICOM_STORE_OPEN_MAXIMUM_ALLOWED)
    found = None
    try:
        for cert in store.Certificates:
            try:
                if thumbprint:
                    if getattr(cert, "Thumbprint", "").lower() == thumbprint.lower():
                        found = cert
                        break
                else:
                    found = cert
                    break
            except Exception:
                continue
    finally:
        store.Close()
    pythoncom.CoUninitialize()
    return found

def sign_data(cert, base64_content: str, b_detached: bool = False) -> str:
    """
    Подписывает данные из base64-строки CAdES-BES подписью.
    base64_content - base64-строка данных для подписи.
    b_detached - True для отсоединенной (detached), False для присоединенной (attached).
    Возвращает подпись в base64 (ASCII).
    """
    pythoncom.CoInitialize()
    signer = Dispatch("CAdESCOM.CPSigner")
    signer.Certificate = cert

    oSigningTimeAttr = Dispatch("CAdESCOM.CPAttribute")
    oSigningTimeAttr.Name = CAPICOM_AUTHENTICATED_ATTRIBUTE_SIGNING_TIME
    oSigningTimeAttr.Value = datetime.datetime.now()
    signer.AuthenticatedAttributes2.Add(oSigningTimeAttr)

    signed_data = Dispatch("CAdESCOM.CadesSignedData")
    signed_data.ContentEncoding = CADESCOM_BASE64_TO_BINARY
    signed_data.Content = base64_content
    signature = signed_data.SignCades(signer, CADES_BES, b_detached, CAPICOM_ENCODE_BASE64)

    if isinstance(signature, bytes):
        signature = signature.decode("ascii", errors="ignore")
    pythoncom.CoUninitialize()
    return signature.replace("\r", "").replace("\n", "")

def post_with_winhttp(url, payload, headers=None):
    win_http = Dispatch('WinHTTP.WinHTTPRequest.5.1')
    win_http.Open("POST", url, False)
    win_http.SetRequestHeader("User-Agent", "Mozilla/5.0")
    win_http.SetRequestHeader("Accept", "application/json, text/plain, */*")
    win_http.SetRequestHeader("Content-Type", "application/json; charset=utf-8")
    if headers:
        for k, v in headers.items():
            win_http.SetRequestHeader(k, v)
    win_http.Send(json.dumps(payload))
    win_http.WaitForResponse()
    status = win_http.Status
    response_text = win_http.ResponseText
    all_headers = win_http.GetAllResponseHeaders()
    if status != 200:
        raise Exception(f"WinHTTP POST failed: Status {status} - {response_text}")
    pythoncom.CoUninitialize()
    return status, response_text, all_headers


# ---------------- Refresh OMS token ----------------
def refresh_oms_token(session: requests.Session, cert, organization_id: str) -> bool:
    logger.info("Обновление токена OMS...")
    url_auth = f"{BASE}/api/v1/crpt/auth?organizationId={organization_id}"

    try:
        resp_get = session.get(url_auth, timeout=15)
        resp_get.raise_for_status()
        challenges = resp_get.json()
        logger.debug(f"Ответ /crpt/auth GET: {json.dumps(challenges, indent=2)}")
        if not isinstance(challenges, list):
            logger.error(f"[ERR] Некорректный формат challenges: {challenges}")
            return False
    except Exception as e:
        logger.error(f"[ERR] GET challenges для OMS: {e}")
        return False

    payload = []
    for ch in challenges:
        if ch['productGroup'] in ['oms', 'trueApi']:
            try:
                sig = sign_data(cert, ch["base64Data"], b_detached=False)  # attached для auth
                payload.append({
                    "uuid": ch["uuid"],
                    "productGroup": ch["productGroup"],
                    "base64Data": sig  # trueApi/oms используют одно поле
                })
            except Exception as e:
                logger.error(f"Подпись challenge для {ch['productGroup']} (uuid={ch['uuid']}): {e}")
                return False


    if not payload:
        logger.error("Нет challenge для OMS в ответе")
        return False

    try:
        cookies_dict = session.cookies.get_dict()
        cookie_str = "; ".join([f"{k}={v}" for k, v in cookies_dict.items()]) if cookies_dict else ""
        custom_headers = {"Cookie": cookie_str} if cookie_str else None
        status, resp_text, all_headers = post_with_winhttp(url_auth, payload, headers=custom_headers)

        # Обновление cookies в session из Set-Cookie
        set_cookie_lines = [line.strip()[len("Set-Cookie:"):].strip() for line in all_headers.splitlines() if line.strip().startswith("Set-Cookie:")]
        if set_cookie_lines:
            temp_resp = requests.Response()
            temp_resp.headers['Set-Cookie'] = set_cookie_lines
            temp_resp.status_code = status
            temp_resp.url = url_auth
            session.cookies.update(temp_resp.cookies)

        logger.info(f"Токен OMS обновлён успешно. Ответ: {resp_text}")
        return True
    except Exception as e:
        logger.error(f"POST signed challenges для OMS: {e}")
        return False

# ---------------- API flows ----------------
def try_single_post(session: requests.Session, document_number: str,
                    product_group: str, release_method_type: str,
                    positions: list[dict],
                    filling_method: str = "productsCatalog", thumbprint: str | None = None) -> dict | None:

    signed_orders_payload: list[dict] = []

    logger.info(f"Создание документа: {document_number}")
    url_create = f"{BASE}/api/v1/codes-order?warehouseId={WAREHOUSE_ID}"
    body = {
        "documentNumber": document_number,
        "comment": "",
        "productGroup": product_group,
        "releaseMethodType": release_method_type,
        "fillingMethod": filling_method,
        "cisType": CIS_TYPE,
        "positions": [
            {
                "gtin": p["gtin"],
                "name": p.get("name", ""),
                "tnvedCode": p.get("tnvedCode", ""),
                "quantity": p.get("quantity", 1),
                "certificateDocument": None,
                "setsGtinUnits": []
            } for p in positions
        ]
    }

    try:
        resp = session.post(url_create, json=body, timeout=30)
        resp.raise_for_status()
        created = resp.json()
        document_id = created.get("id") if isinstance(created, dict) else str(created).strip('"')
    except Exception as e:
        logger.error(f"Создание документа {document_number}: {e}")
        return None

    logger.info(f"Документ создан: {document_id}")

    # проверка доступности
    try:
        resp = session.get(f"{BASE}/api/v1/codes-order/{document_id}/availability-status", timeout=15)
        resp.raise_for_status()
        status = resp.text.strip('"')
        if status != "available":
            logger.warning(f"Документ {document_number} недоступен: {status}")
            return None
    except Exception as e:
        logger.error(f"Проверка доступности документа {document_number}: {e}")
        return None

    # проверка сертификата
    if thumbprint:
        try:
            resp = session.get(f"{BASE}/api/v1/organizations/{ORGANIZATION_ID}/employees/has-certificate?thumbprint={thumbprint}", timeout=15)
            resp.raise_for_status()
            if not resp.json():
                logger.error("Сертификат не зарегистрирован в организации")
                return None
        except Exception as e:
            logger.error(f"Проверка сертификата: {e}")
            return None

    # обновление токена OMS
    cert = find_certificate_by_thumbprint(thumbprint)
    if not cert:
        logger.error(f"Сертификат для подписи не найден (thumbprint={thumbprint})")
        return None

    # получение orders-for-sign
    try:
        resp = session.get(f"{BASE}/api/v1/codes-order/{document_id}/orders-for-sign", timeout=15)
        resp.raise_for_status()
        orders_to_sign = resp.json()
        if not isinstance(orders_to_sign, list):
            logger.error(f"Некорректный формат orders_for_sign: {orders_to_sign}")
            return None
    except Exception as e:
        logger.error(f"Получение данных для подписи: {e}")
        return None

    # подпись каждого order
    for o in orders_to_sign:
        oid = o["id"]
        b64content = o["base64Content"]
        logger.info(f"Подписываем order id={oid} (base64Content length={len(b64content)})")
        try:
            signature_b64 = sign_data(cert, b64content, b_detached=True)  # detached для orders
            signed_orders_payload.append({"id": oid, "base64Content": signature_b64})
        except Exception as e:
            logger.error(f"Ошибка подписи order {oid}: {e}")
            return None

    # отправка документа
    try:
        if not refresh_oms_token(session, cert, str(ORGANIZATION_ID)):
            logger.error("Не удалось обновить токен OMS")
            return None

        send_url = f"{BASE}/api/v1/codes-order/{document_id}/send"
        payload = {"signedOrders": signed_orders_payload}
        r_send = session.post(send_url, json=payload, timeout=30)
        r_send.raise_for_status()
        logger.info("Отправка прошла успешно (detached signature)")
    except Exception as e:
        logger.info(f"Отправка документа {document_number}: {e}")
        return None

    # финальный статус
    try:
        r_fin = session.get(f"{BASE}/api/v1/codes-order/{document_id}", timeout=15)
        r_fin.raise_for_status()
        doc = r_fin.json()
        logger.info(f"Финальный статус документа: {doc.get('status')}")
        return doc
    except Exception as e:
        logger.error(f"Получение финального статуса: {e}")
        return None

# ---------------- Download PDF ----------------
def get_with_winhttp(url: str, headers: dict | None = None):
    """
    Выполнить GET через WinHTTP и вернуть (status, response_bytes, all_headers)
    """
    win_http = Dispatch('WinHTTP.WinHTTPRequest.5.1')
    win_http.Open("GET", url, False)
    # стандартные заголовки
    win_http.SetRequestHeader("User-Agent", "Mozilla/5.0")
    if headers:
        for k, v in headers.items():
            # не пересоздаём User-Agent
            if k.lower() == "user-agent":
                continue
            win_http.SetRequestHeader(k, v)
    win_http.Send()
    win_http.WaitForResponse()
    status = int(win_http.Status)
    # ResponseBody в COM возвращает бинарные данные
    try:
        body = win_http.ResponseBody  # binary
    except Exception:
        body = None
    all_headers = win_http.GetAllResponseHeaders()
    return status, body, all_headers


def check_order_status(session: requests.Session, document_id: str) -> str:
    """
    Быстрая проверка статуса заказа без ожидания
    """
    try:
        resp_status = session.get(f"{BASE}/api/v1/codes-order/{document_id}", timeout=15)
        resp_status.raise_for_status()
        doc = resp_status.json()
        return doc.get("status", "unknown")
    except Exception as e:
        logger.error(f"Ошибка проверки статуса заказа {document_id}: {e}")
        return "error"
    
    
def download_codes_pdf_and_convert(session: requests.Session, document_id: str, order_name: str) -> Optional[Tuple[Optional[str], Optional[str], Optional[str]]]:
    """
    Скачивает PDF, CSV и XLS (если доступны) для заказа document_id и сохраняет их в папку:
      Desktop / "pdf-коды км" / <safe_order_name>/
    Файлы сохраняются с общей базой имени, производной от order_name.
    Возвращает кортеж (pdf_path, csv_path, xls_path). Если файл не скачан — соответствующий элемент = None.
    """
    logger.info(f"Начало скачивания PDF/CSV/XLS для заказа {document_id} ({order_name!r})")

    from urllib.parse import urljoin

    def make_full_url(url: str) -> str:
        if not url:
            return ""
        if url.startswith("http://") or url.startswith("https://"):
            return url
        return urljoin(BASE, url)

    # 1) дождаться статуса released (polling)
    max_attempts = 10  # 5 минут (10 * 30s)
    attempt = 0
    status = None
    while attempt < max_attempts:
        try:
            resp_status = session.get(f"{BASE}/api/v1/codes-order/{document_id}", timeout=15)
            resp_status.raise_for_status()
            doc = resp_status.json()
            status = doc.get("status")
            logger.info(f"Статус заказа {document_id}: {status}")
            if status == "released":
                break
            time.sleep(30)
            attempt += 1
        except Exception as e:
            logger.error(f"Ошибка проверки статуса заказа {document_id}: {e}", exc_info=True)
            return None

    if status != "released":
        logger.error(f"Заказ {document_id} не перешёл в 'released' за {max_attempts * 30} сек")
        return None

    # Подготовка каталога для сохранения: Desktop / "pdf-коды км" / <safe_order_name>
    try:
        desktop = os.path.join(os.path.expanduser("~"), "Desktop")
        parent_dir = os.path.join(desktop, "pdf-коды км")
        # безопасное имя папки
        safe_order_name = "".join(c for c in (order_name or document_id) if c.isalnum() or c in " -_").strip()
        if not safe_order_name:
            safe_order_name = document_id
        safe_order_name = safe_order_name[:120]
        target_dir = os.path.join(parent_dir, safe_order_name)
        os.makedirs(target_dir, exist_ok=True)
        # базовое безопасное имя файла
        safe_base = safe_order_name[:100]
    except Exception as e:
        logger.error(f"Ошибка при подготовке пути сохранения: {e}", exc_info=True)
        return None

    # Заголовки (включая куки из session)
    cookies_dict = session.cookies.get_dict()
    cookie_str = "; ".join([f"{k}={v}" for k, v in cookies_dict.items()]) if cookies_dict else ""
    headers = {"User-Agent": "Mozilla/5.0", "Referer": BASE}
    if cookie_str:
        headers["Cookie"] = cookie_str

    pdf_path = None
    csv_path = None
    xls_path = None

    # ---------------- PDF export ----------------
    try:
        # получить templateId для size "30x20"
        resp_templates = session.get(f"{BASE}/api/v1/print-templates?organizationId={ORGANIZATION_ID}&formTypes=codesOrder", timeout=15)
        resp_templates.raise_for_status()
        templates = resp_templates.json()
        template_id = None
        for t in templates:
            if t.get("size") == "30x20":
                template_id = t.get("id")
                break
        if template_id:
            export_url = f"{BASE}/api/v1/codes-order/{document_id}/export/pdf?splitByGtins=false&templateId={template_id}"
            resp_export = session.post(export_url, timeout=30)
            resp_export.raise_for_status()
            export_data = resp_export.json()
            result_id = export_data.get("resultId")
            logger.info(f"PDF export started for {document_id}, resultId: {result_id}")

            # polling результата
            file_url = None
            attempts_pdf = 0
            while attempts_pdf < 12:  # ~2 минуты
                resp_result = session.get(f"{BASE}/api/v1/codes-order/{document_id}/export/pdf/{result_id}", timeout=15)
                resp_result.raise_for_status()
                result_data = resp_result.json()
                if result_data.get("status") == "success":
                    file_infos = result_data.get("fileInfos", [])
                    if file_infos:
                        file_url = file_infos[0].get("fileUrl") or file_infos[0].get("fileUrlAbsolute") or None
                    break
                time.sleep(10)
                attempts_pdf += 1

            if file_url:
                full_file_url = make_full_url(file_url)
                safe_pdf_name = f"{safe_base}.pdf"
                pdf_path = os.path.join(target_dir, safe_pdf_name)
                try:
                    logger.debug(f"Downloading PDF from {full_file_url}")
                    r = session.get(full_file_url, timeout=60, headers=headers, stream=True, allow_redirects=True)
                    r.raise_for_status()
                    with open(pdf_path, "wb") as fh:
                        for chunk in r.iter_content(chunk_size=8192):
                            if chunk:
                                fh.write(chunk)
                    logger.info(f"PDF сохранён: {pdf_path}")
                except Exception as e:
                    logger.error(f"Ошибка скачивания PDF (requests) {full_file_url}: {e}", exc_info=True)
                    pdf_path = None
            else:
                logger.warning(f"PDF export для {document_id} завершился без fileUrl")
        else:
            logger.warning("Шаблон '30x20' для PDF не найден — пропускаем PDF экспорт")
    except Exception as e:
        logger.exception(f"Ошибка в PDF-части для {document_id}: {e}")
        pdf_path = None

    # ---------------- CSV export ----------------
    try:
        export_csv_url = f"{BASE}/api/v1/codes-order/{document_id}/export/csv?splitByGtins=false"
        resp_csv_export = session.post(export_csv_url, timeout=30)
        resp_csv_export.raise_for_status()
        csv_export_data = resp_csv_export.json()
        csv_result_id = csv_export_data.get("resultId")
        logger.info(f"CSV export started for {document_id}, resultId: {csv_result_id}")

        # polling CSV result
        file_infos = None
        attempts_csv = 0
        while attempts_csv < 30:  # до ~5 минут (30 * 10s)
            resp_csv_status = session.get(f"{BASE}/api/v1/codes-order/{document_id}/export/csv/{csv_result_id}", timeout=15)
            resp_csv_status.raise_for_status()
            csv_status_data = resp_csv_status.json()
            status_csv = csv_status_data.get("status")
            logger.info(f"CSV export status for {document_id}: {status_csv}")
            if status_csv == "success":
                file_infos = csv_status_data.get("fileInfos", [])
                break
            time.sleep(10)
            attempts_csv += 1

        if file_infos:
            finfo = file_infos[0]
            file_id = finfo.get("fileId")
            download_csv_url = f"{BASE}/api/v1/codes-order/{document_id}/export/csv/{csv_result_id}/download/{file_id}"
            # защитить имя файла
            safe_csv_name = f"{order_name}_csv.csv"
            if any(ch in safe_csv_name for ch in ("/", "\\", ":", "*", "?", "\"", "<", ">", "|")):
                safe_csv_name = f"{safe_base}.csv"
            csv_path = os.path.join(target_dir, safe_csv_name)
            try:
                logger.debug(f"Downloading CSV from {download_csv_url}")
                r = session.get(download_csv_url, timeout=60, headers=headers, stream=True, allow_redirects=True)
                r.raise_for_status()
                with open(csv_path, "wb") as fh:
                    for chunk in r.iter_content(chunk_size=8192):
                        if chunk:
                            fh.write(chunk)
                logger.info(f"CSV сохранён: {csv_path}")
            except Exception as e:
                logger.error(f"Ошибка скачивания CSV (requests) {download_csv_url}: {e}", exc_info=True)
                csv_path = None
        else:
            logger.warning(f"CSV export для {document_id} не вернул fileInfos в пределах таймаута")
    except Exception as e:
        logger.exception(f"Ошибка в CSV-части для {document_id}: {e}")
        csv_path = None

    # ---------------- XLS export ----------------
    try:
        export_xls_url = f"{BASE}/api/v1/codes-order/{document_id}/export/xls?splitByGtins=false"
        resp_xls_export = session.post(export_xls_url, timeout=30)
        resp_xls_export.raise_for_status()
        xls_export_data = resp_xls_export.json()
        xls_result_id = xls_export_data.get("resultId")
        logger.info(f"XLS export started for {document_id}, resultId: {xls_result_id}")

        # polling XLS result
        file_infos = None
        attempts_xls = 0
        while attempts_xls < 30:  # до ~5 минут
            resp_xls_status = session.get(f"{BASE}/api/v1/codes-order/{document_id}/export/xls/{xls_result_id}", timeout=15)
            resp_xls_status.raise_for_status()
            xls_status_data = resp_xls_status.json()
            status_xls = xls_status_data.get("status")
            logger.info(f"XLS export status for {document_id}: {status_xls}")
            if status_xls == "success":
                file_infos = xls_status_data.get("fileInfos", [])
                break
            time.sleep(10)
            attempts_xls += 1

        if file_infos:
            finfo = file_infos[0]
            file_id = finfo.get("fileId")
            download_xls_url = f"{BASE}/api/v1/codes-order/{document_id}/export/xls/{xls_result_id}/download/{file_id}"
            safe_xls_name = f"{order_name}.xls"
            if any(ch in safe_xls_name for ch in ("/", "\\", ":", "*", "?", "\"", "<", ">", "|")):
                safe_xls_name = f"{safe_base}.xls"
            xls_path = os.path.join(target_dir, safe_xls_name)
            try:
                logger.debug(f"Downloading XLS from {download_xls_url}")
                r = session.get(download_xls_url, timeout=60, headers=headers, stream=True, allow_redirects=True)
                r.raise_for_status()
                with open(xls_path, "wb") as fh:
                    for chunk in r.iter_content(chunk_size=8192):
                        if chunk:
                            fh.write(chunk)
                logger.info(f"XLS сохранён: {xls_path}")
            except Exception as e:
                logger.error(f"Ошибка скачивания XLS (requests) {download_xls_url}: {e}", exc_info=True)
                xls_path = None
        else:
            logger.warning(f"XLS export для {document_id} не вернул fileInfos в пределах таймаута")
    except Exception as e:
        logger.exception(f"Ошибка в XLS-части для {document_id}: {e}")
        xls_path = None

    # Вернуть кортеж путей (возможно некоторые элементы None)
    return pdf_path, csv_path, xls_path




def perform_introduction_from_order(
    session: requests.Session,
    codes_order_id: str,
    organization_id: str = ORGANIZATION_ID,
    thumbprint: Optional[str] = None,
    # опциональные поля для PATCH /production (если нужно обновить метаданные)
    production_patch: Optional[Dict[str, Any]] = None,
    # автоперезапросы / таймауты
    check_poll_interval: int = 5,      # сек для /codes-checking polling (короткий интервал)
    check_poll_attempts: int = 24,     # сколько попыток (24*5=120s default)
    gen_poll_interval: int = 2,        # сек для ожидания generate/send результатов
    gen_poll_attempts: int = 30        # сколько попыток для generate (примерно 60s)
) -> Tuple[bool, Dict[str, Any]]:
    """
    Выполняет ввод в оборот (codes introduction) для указанного codes_order_id.
    Возвращает (ok: bool, result: dict) где result содержит поля:
      - introduction_id
      - created_introduction (ответ GET после создания)
      - check_status (последний ответ /codes-checking)
      - production (последний GET /production)
      - generate_items (raw items from generate-multiple)
      - send_response (ответ от send-multiple)
      - final_introduction, final_check (последние GET ответы)
      - errors (список сообщений об ошибках)
    """
    result: Dict[str, Any] = {"errors": []}
    try:
        # 1) create-from-codes-order
        try:
            addr = socket.getaddrinfo('mk.kontur.ru', 443)
            logger.info(f"DNS успешно разрешён: {addr}")
            url_create = f"{BASE}/api/v1/codes-introduction/create-from-codes-order/{codes_order_id}?isImportFts=false&isAccompanyingDocumentNeeds=false"
            extra_kwargs = {}  # Без verify=False
        except socket.gaierror as dns_err:
            logger.error(f"DNS-разрешение провалено: {dns_err}. Используем workaround с IP.")
            # Workaround: Используем IP вместо домена
            ip = '46.17.200.242'
            original_base = BASE.replace('mk.kontur.ru', ip)
            url_create = f"{original_base}/api/v1/codes-introduction/create-from-codes-order/{codes_order_id}?isImportFts=false&isAccompanyingDocumentNeeds=false"
            extra_kwargs = {'verify': False}  # Отключаем проверку SSL (небезопасно!)
        # Добавляем заголовок Host для workaround (если используем IP)
        headers = {'Host': 'mk.kontur.ru'} if 'ip' in locals() else {}
        
        r = session.post(url_create, headers=headers, timeout=30, **extra_kwargs)
        logger.info(f"Ответ сервера: {r.text}")
        r.raise_for_status()
        intro_id = r.text.strip().strip('"')
        result["introduction_id"] = intro_id
        logger.info("Создана заявка ввода в оборот: %s", intro_id)
        
        # 2) initial GET introduction
        r_intro = session.get(f"{BASE}/api/v1/codes-introduction/{intro_id}", timeout=15)
        r_intro.raise_for_status()
        result["created_introduction"] = r_intro.json()

        # 3) poll codes-checking until status indicates no errors
        check_ok = False
        attempts = 0
        last_check = None
        while attempts < check_poll_attempts:
            r_check = session.get(f"{BASE}/api/v1/codes-checking/{intro_id}", timeout=15)
            if r_check.status_code == 200:
                last_check = r_check.json()
                result["check_status_latest"] = last_check
                status = last_check.get("status")
                logger.info("codes-checking status for %s: %s", intro_id, status)
                if status in ("inProgress", "doesNotHaveErrors", "created", "checked", "noErrors"):  # возможные статусы
                    check_ok = True
                    break
            else:
                logger.debug("codes-checking returned non-200: %s", r_check.status_code)
            attempts += 1
            time.sleep(check_poll_interval)

        if not check_ok:
            msg = f"codes-checking для {intro_id} не перешёл в OK-статус после {check_poll_attempts} попыток"
            logger.warning(msg)
            result["errors"].append(msg)
            # продолжим, возможно всё равно можно отправить (но лучше вернуть ошибку)
            # return False, result

        # 4) GET production (получаем структуру production)
        try:
            r_prod = session.get(f"{BASE}/api/v1/codes-introduction/{intro_id}/production", timeout=15)
            r_prod.raise_for_status()
            result["production"] = r_prod.json()
        except Exception as e:
            logger.warning("Не удалось получить /production: %s", e)
            result["errors"].append(f"production GET error: {e}")

        # 5) Если переданы поля для PATCH /production — применим
        if production_patch:
            try:
                patch_url = f"{BASE}/api/v1/codes-introduction/{intro_id}/production"
                logger.info("PATCH production %s", patch_url)
                r_patch = session.patch(patch_url, json=production_patch, timeout=30)
                r_patch.raise_for_status()
                result["production_patch_response"] = r_patch.json() if r_patch.content else {"status": "ok"}
                # обновим production
                r_prod2 = session.get(f"{BASE}/api/v1/codes-introduction/{intro_id}/production", timeout=15)
                r_prod2.raise_for_status()
                result["production_after_patch"] = r_prod2.json()
            except Exception as e:
                logger.exception("Ошибка PATCH production: %s", e)
                result["errors"].append(f"production PATCH error: {e}")

        # 6) попытка автозаполнения позиций (autocomplete) — не обязателен, но делаем если нужно
        try:
            auto_url = f"{BASE}/api/v1/codes-introduction/{intro_id}/positions/autocomplete"
            logger.info("POST positions/autocomplete ...")
            r_auto = session.post(auto_url, timeout=30)
            # многие реализации возвращают 204/200/empty; не обязательно требовать body
            if r_auto.status_code not in (200, 204):
                logger.debug("autocomplete returned status %s", r_auto.status_code)
            result["autocomplete_status"] = r_auto.status_code
        except Exception as e:
            logger.warning("autocomplete failed: %s", e)
            result["errors"].append(f"autocomplete error: {e}")

        # 7) проверить наличие сертификата у сотрудника в организации
        try:
            if thumbprint:
                r_cert_check = session.get(f"{BASE}/api/v1/organizations/{organization_id}/employees/has-certificate?thumbprint={thumbprint}", timeout=15)
                r_cert_check.raise_for_status()
                has_cert = bool(r_cert_check.json())
                result["has_certificate"] = has_cert
                if not has_cert:
                    msg = "Сертификат с указанным thumbprint не зарегистрирован в организации"
                    logger.error(msg)
                    result["errors"].append(msg)
                    return False, result
        except Exception as e:
            logger.exception("Ошибка проверки сертификата: %s", e)
            result["errors"].append(f"cert check error: {e}")
            return False, result

        # 8) GET generate-multiple -> получим массив объектов base64Content
        try:
            gen_url = f"{BASE}/api/v1/codes-introduction/{intro_id}/generate-multiple"
            r_gen = session.get(gen_url, timeout=30)
            logger.info(f"Ответ сервера при вводе в оборот: {r_gen.json()}")
            r_gen.raise_for_status()
            gen_items = r_gen.json()
            result["generate_items_raw"] = gen_items
            if not isinstance(gen_items, list) or not gen_items:
                msg = "generate-multiple вернул пустой список"
                logger.error(msg)
                result["errors"].append(msg)
                return False, result
        except Exception as e:
            logger.exception("Ошибка generate-multiple: %s", e)
            result["errors"].append(f"generate-multiple error: {e}")
            return False, result

        # 9) подпись каждого base64Content
        cert = find_certificate_by_thumbprint(thumbprint)
        if not cert:
            msg = f"Сертификат для подписи не найден (thumbprint={thumbprint})"
            logger.error(msg)
            result["errors"].append(msg)
            return False, result

        signed_payloads: List[Dict[str, str]] = []
        for item in gen_items:
            docid = item.get("documentId") or item.get("documentId")
            b64 = item.get("base64Content")
            if not b64:
                result["errors"].append(f"item {docid} missing base64Content")
                continue
            # попробуем сначала attached (b_detached=False), если ошибка — detached True
            sig = None
            try:
                sig = sign_data(cert, b64, b_detached=True)
                # sign_data в твоей реализации возвращает строку (или кортеж?), убедимся что строка:
                if isinstance(sig, tuple):
                    sig = sig[0]
            except Exception as e:
                logger.warning("Attached sign failed for %s: %s — попробуем detached", docid, e)
                try:
                    sig = sign_data(cert, b64, b_detached=True)
                    if isinstance(sig, tuple):
                        sig = sig[0]
                except Exception as e2:
                    logger.exception("Не удалось подписать %s ни attached ни detached: %s", docid, e2)
                    result["errors"].append(f"sign failed for {docid}: {e2}")
                    continue
            if not sig:
                result["errors"].append(f"signature empty for {docid}")
                continue
            signed_payloads.append({"documentId": docid, "signedContent": sig})

        if not signed_payloads:
            msg = "Нет подписанных документов для отправки"
            logger.error(msg)
            result["errors"].append(msg)
            return False, result

        result["signed_payloads_preview"] = [{"documentId": p["documentId"], "signed_len": len(p["signedContent"])} for p in signed_payloads]


        # 11) отправка send-multiple
        try:
            send_url = f"{BASE}/api/v1/codes-introduction/{intro_id}/send-multiple"
            logger.info("Отправка send-multiple ...")
            r_send = session.post(send_url, json=signed_payloads, timeout=30)
            r_send.raise_for_status()
            # API возвращает массив / подтверждение — сохраним
            try:
                result["send_response"] = r_send.json()
            except Exception:
                result["send_response"] = r_send.text
        except Exception as e:
            logger.exception("Ошибка send-multiple: %s", e)
            result["errors"].append(f"send-multiple error: {e}")
            return False, result

        # 12) финальные GET'ы
        try:
            r_final_intro = session.get(f"{BASE}/api/v1/codes-introduction/{intro_id}", timeout=15)
            r_final_intro.raise_for_status()
            result["final_introduction"] = r_final_intro.json()
        except Exception as e:
            logger.warning("final introduction GET failed: %s", e)

        try:
            r_final_check = session.get(f"{BASE}/api/v1/codes-checking/{intro_id}", timeout=15)
            r_final_check.raise_for_status()
            result["final_check"] = r_final_check.json()
        except Exception as e:
            logger.warning("final checking GET failed: %s", e)

        # Всё успешно
        ok = not bool(result["errors"])
        return ok, result
        
    except Exception as e:
        result["errors"].append(str(e))
        logger.error(f"Ошибка в perform_introduction_from_order: {e}")
        return False, result


def perform_introduction_from_order_tsd(
    session: requests.Session,
    codes_order_id: str,
    positions_data: List[Dict[str, str]],  # Список {'name': str, 'gtin': str} для позиций в XLS
    organization_id: str = ORGANIZATION_ID,
    warehouse_id: str = WAREHOUSE_ID,  # Фиксированный склад
    thumbprint: Optional[str] = None,
    production_patch: Optional[Dict[str, Any]] = None,  # Только productionDate, expirationDate, batchNumber из GUI
    check_poll_interval: int = 5,
    check_poll_attempts: int = 24,
    tsd_poll_interval: int = 5,  # Для polling после send-to-tsd
    tsd_poll_attempts: int = 20
) -> Tuple[bool, Dict[str, Any]]:
    """
    Создаёт задание ввода в оборот через ТСД для codes_order_id.
    Автоматически генерирует XLS на основе positions_data.
    Возвращает (ok: bool, result: dict) с полями: introduction_id, errors, final_status, etc.
    """
    result: Dict[str, Any] = {"errors": []}
    try:
        # Автоматическая генерация XLS
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_file:
            wb = openpyxl.Workbook()
            ws = wb.active
            # Заголовки (опционально, но в логах без — просто данные)
            # ws['A1'] = "Name"
            # ws['B1'] = "GTIN"
            row_num = 1
            for pos in positions_data:
                ws.cell(row=row_num, column=1, value=pos.get('name', ''))
                ws.cell(row=row_num, column=2, value=pos.get('gtin', ''))
                row_num += 1
            wb.save(tmp_file.name)
            file_path = tmp_file.name

        # Теперь используем file_path как раньше
        with open(file_path, "rb") as f:
            file_content = f.read()
            base64_content = base64.b64encode(file_content).decode('utf-8')

        # 1. POST /codes-introduction?warehouseId — Создать документ
        url_create = f"{BASE}/api/v1/codes-introduction?warehouseId={warehouse_id}"
        logger.info("Создаём ввод в оборот для ТСД...")
        r = session.post(url_create, timeout=30)
        r.raise_for_status()
        intro_id = r.text.strip().strip('"')
        result["introduction_id"] = intro_id
        logger.info("Создана заявка: %s", intro_id)

        # 2. GET /codes-introduction/{id} — Проверить создание
        r_intro = session.get(f"{BASE}/api/v1/codes-introduction/{intro_id}", timeout=15)
        r_intro.raise_for_status()
        result["created_introduction"] = r_intro.json()

        # 3. POST /production — Обновить метаданные (с fillingMethod="tsd")
        patch_url = f"{BASE}/api/v1/codes-introduction/{intro_id}/production"
        full_patch = {
            "documentNumber": production_patch.get("documentNumber", "NO_NAME"),  # Из item/order_name
            "productionType": "ownProduction",
            "warehouseId": warehouse_id,
            "expirationType": "milkMoreThan72",
            "containsUtilisationReport": "true",
            "usageType": "verified",
            "cisType": "unit",
            "fillingMethod": "tsd",
            "isAutocompletePositionsDataNeeded": "true",
            "productsHasSameDates": "true",
            "productionDate": production_patch.get("productionDate"),
            "expirationDate": production_patch.get("expirationDate"),  # Или рассчитать +5 лет
            "batchNumber": production_patch.get("batchNumber"),
            "TnvedCode": production_patch.get("TnvedCode", "")  # Если есть
        }
        logger.info("PATCH production для ТСД...")
        r_patch = session.post(patch_url, json=full_patch, timeout=30)
        r_patch.raise_for_status()
        result["production_patch_response"] = r_patch.json() if r_patch.content else {"status": "ok"}

        # 4. Upload XLS и parse
        # Инициация upload
        init_url = f"{BASE}/drive/v1/contents/js/initPartialUpload-v2?client=js_v0&UploadId={codes_order_id}&source=blob"  # Используем codes_order_id как UploadId или генерировать
        init_payload = {
            "Name": f"srv/upload/uid/{organization_id}/auto_generated_{codes_order_id}.xlsx",
            "Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"  # Для XLSX
        }
        r_init = session.post(init_url, json=init_payload, timeout=30)
        r_init.raise_for_status()
        init_resp = r_init.json()
        session_id = init_resp["UploadSessionId"]
        parts_uri = init_resp["PartsUploadUri"]

        # PUT parts (base64 upload)
        parts_url = f"{parts_uri}&client=js_v0&source=blob&te=base64"
        r_parts = session.put(parts_url, data=base64_content, timeout=60)
        r_parts.raise_for_status()

        # Parse importer
        file_uri = init_resp["Name"]  # Из init
        parse_url = f"{BASE}/import/v1/api/Importer/Parse?fileUri={file_uri}&disableDataBoundsDetection=true&extension=xlsx"
        r_parse = session.get(parse_url, timeout=30)
        r_parse.raise_for_status()
        parse_data = r_parse.json()["data"]  # [[name, gtin], ...]

        # 5. POST /positions — Отправить распарсеные данные
        positions_url = f"{BASE}/api/v1/codes-introduction/{intro_id}/positions"
        positions_payload = {
            "rows": [
                {
                    "name": row[0],
                    "gtin": row[1],
                    "tnvedCode": production_patch.get("TnvedCode", ""),  # Если нужно
                    "certificateDocumentNumber": "",
                    "certificateDocumentDate": "",
                    "costInKopecksWithVat": 0,
                    "exciseInKopecks": 0
                } for row in parse_data
            ]
        }
        r_positions = session.post(positions_url, json=positions_payload, timeout=30)
        r_positions.raise_for_status()
        result["positions_response"] = r_positions.json() if r_positions.content else {"status": "ok"}

        # 6. GET /production — Проверить обновления
        r_prod = session.get(f"{BASE}/api/v1/codes-introduction/{intro_id}/production", timeout=15)
        r_prod.raise_for_status()
        result["production_final"] = r_prod.json()

        # 7. POST /send-to-tsd — Отправить задание на ТСД
        tsd_url = f"{BASE}/api/v1/codes-introduction/{intro_id}/send-to-tsd"
        r_tsd = session.post(tsd_url, timeout=30)
        r_tsd.raise_for_status()
        result["tsd_response"] = r_tsd.json() if r_tsd.content else {"status": "ok"}

        # Опциональный polling статуса после send-to-tsd
        tsd_ok = False
        attempts = 0
        while attempts < tsd_poll_attempts:
            r_final = session.get(f"{BASE}/api/v1/codes-introduction/{intro_id}", timeout=15)
            if r_final.status_code == 200:
                final_data = r_final.json()
                result["final_introduction"] = final_data
                status = final_data.get("documentStatus")
                if status in ("sent_to_tsd", "completed"):  # Предполагаемые статусы
                    tsd_ok = True
                    break
            attempts += 1
            time.sleep(tsd_poll_interval)

        if not tsd_ok:
            result["errors"].append("Не удалось подтвердить отправку на ТСД")

        # Удаляем temp файл
        os.unlink(file_path)

        ok = not bool(result["errors"])
        return ok, result

    except Exception as e:
        logger.exception("Ошибка в perform_introduction_from_order_tsd: %s", e)
        result["errors"].append(str(e))
        if 'file_path' in locals():
            os.unlink(file_path)
        return False, result